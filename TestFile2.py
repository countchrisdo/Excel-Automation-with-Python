from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
print('TestFile2.py Running')

### Importing workbook and setting up worksheet
wb = load_workbook('grades.xlsx')
ws = wb.active

### looping through row 1-10 / Rows are Numbers on side / Columns are Letters on Top
for row in range(1, 11):
    for col in range(1,5):
        #imported function that translates numbers and letters in Excel
        char = get_column_letter(col)
        #printing the value of each cell in the loop
        print(ws[char + str(row)].value)
        ws[char + str(row)] = char + str(row)


wb.save('tim.xlsx')
print('TestFile2.py Ran')

