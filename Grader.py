from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
print('working.py Running')

#Creating Data
data = {
	"Joe": {
		"Math": 65,
		"Science": 78,
		"English": 98,
		"Gym": 89
	},
	"Bill": {
		"Math": 55,
		"Science": 72,
		"English": 87,
		"Gym": 95
	},
	"Tim": {
		"Math": 100,
		"Science": 45,
		"english": 75,
		"Gym": 92
	},
	"Sally": {
		"Math": 30,
		"Science": 25,
		"English": 45,
		"Gym": 100
	},
	"Jane": {
		"Math": 100,
		"Science": 100,
		"English": 100,
		"Gym": 60
	}
}

wb = Workbook()
ws = wb.active
ws.title = "Grades"

headings = ['Name'] + list(data['Joe'].keys())
ws.append(headings)

### Inserting Data into sheet
for person in data:
    grades = list(data[person].values())
    ws.append([person] + grades)

### Checking for the amount of data dynamically instead of a fixed number in the loop
### Staring at Col2, then loop through the amount of subjects (4) / then add 2 again to always start at Col2
for col in range(2, len(data['Joe']) + 2):
    char = get_column_letter(col)
    ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

### Styling
for col in range(1, 6):
    ws[get_column_letter(col) + '1'].font = Font(bold=True, color="FF0000")

wb.save('Grades3.xlsx')
print('working.py Ran')

