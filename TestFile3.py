from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
print('TF3.py Running')

### Importing workbook and setting up worksheet
wb = load_workbook('tim.xlsx')
ws = wb.active

### Inserting an empty row at speicifed index
ws.insert_rows(2)
ws.insert_rows(2)
### Deleteing
ws.delete_rows(2)
ws.delete_rows(2)

### inserting Columns
ws.insert_cols(2)
### Deleting Columns
ws.delete_cols(2)

### moving
ws.move_range("C1:D11", rows=2, cols=2)


wb.save('tim.xlsx')
print('TF3.py Ran')

