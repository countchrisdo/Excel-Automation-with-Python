from openpyxl import Workbook, load_workbook

### Importing desired workbook / Must be xlsv / Can use absolute path if file is outside folder
wb = load_workbook('Grades.xlsx')

### Assigns and prints out the active worksheet from this workbook
ws = wb.active
print(ws)

### prints out the Value at A2
print(ws['A2'].value)

#changes the value at A2
ws['A2'].value = "Test"

### Assigns and prints out whatever worksheet I enter
ws = wb['Sheet1']
print(ws)

### creates a new sheet with the given name / Prints are only to check that it works
print(wb.sheetnames)
wb.create_sheet("Test")
print(wb.sheetnames)

### saves your work to a file name you give / This can overwrite a file so be careful 
wb.save('Grades2.xlsx')